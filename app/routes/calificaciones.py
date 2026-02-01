"""
Rutas para gestión de calificaciones
"""

from flask import Blueprint, render_template, request, redirect, url_for, flash, session
from app.models import db, AsignacionCurso, Estudiante, Calificacion
from functools import wraps
from werkzeug.utils import secure_filename
import os
from datetime import datetime
import csv
import openpyxl

calificaciones_bp = Blueprint('calificaciones', __name__)

# Configuración de archivos permitidos
ALLOWED_EXTENSIONS = {'csv', 'xlsx', 'xls'}
UPLOAD_FOLDER = 'uploads'

def allowed_file(filename):
    """Verifico si la extensión del archivo es válida para importación"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Importo el decorador de autenticación del módulo auth
from app.routes.auth import login_requerido

@calificaciones_bp.route('/asignacion/<int:id_asignacion>')
@login_requerido
def listar(id_asignacion):
    """Listar calificaciones de una asignación"""
    asignacion = AsignacionCurso.query.get_or_404(id_asignacion)
    estudiantes = Estudiante.query.filter_by(id_asignacion=id_asignacion).all()
    
    return render_template('calificaciones/listar.html', 
                         asignacion=asignacion,
                         estudiantes=estudiantes)

@calificaciones_bp.route('/nueva/<int:id_asignacion>/<int:id_estudiante>', methods=['GET', 'POST'])
@login_requerido
def nueva(id_asignacion, id_estudiante):
    """Añadir nueva calificación manual"""
    asignacion = AsignacionCurso.query.get_or_404(id_asignacion)
    estudiante = Estudiante.query.get_or_404(id_estudiante)
    
    if request.method == 'POST':
        try:
            # Crear la calificación
            calificacion = Calificacion(
                id_estudiante=id_estudiante,
                nombre_prueba=request.form.get('nombre_prueba', '').strip(),
                tipo_prueba=request.form.get('tipo_prueba', '').strip() or None,
                nota_final=float(request.form.get('nota_final', 0)),
                peso_nota=float(request.form.get('peso_nota', 0)) if request.form.get('peso_nota') else None,
                fecha_prueba=datetime.strptime(request.form.get('fecha_prueba'), '%Y-%m-%d').date() if request.form.get('fecha_prueba') else None,
                elemento_nota=request.form.get('elemento_nota', '').strip() or None,
                fecha_importacion=datetime.now()
            )
            
            db.session.add(calificacion)
            db.session.commit()
            
            flash(f'Calificación añadida correctamente para {estudiante.nombre_alumno}', 'success')
            return redirect(url_for('calificaciones.listar', id_asignacion=id_asignacion))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al guardar la calificación: {str(e)}', 'error')
            return render_template('calificaciones/formulario.html', 
                                 asignacion=asignacion,
                                 estudiante=estudiante)
    
    return render_template('calificaciones/formulario.html', 
                         asignacion=asignacion,
                         estudiante=estudiante)

@calificaciones_bp.route('/importar/<int:id_asignacion>', methods=['GET', 'POST'])
@login_requerido
def importar(id_asignacion):
    """Importar calificaciones desde CSV o Excel"""
    asignacion = AsignacionCurso.query.get_or_404(id_asignacion)
    
    if request.method == 'POST':
        # Verificar que se subió un archivo
        if 'archivo' not in request.files:
            flash('No se seleccionó ningún archivo', 'error')
            return redirect(request.url)
        
        file = request.files['archivo']
        
        if file.filename == '':
            flash('No se seleccionó ningún archivo', 'error')
            return redirect(request.url)
        
        if file and allowed_file(file.filename):
            try:
                # Asegurar que existe la carpeta de uploads
                os.makedirs(UPLOAD_FOLDER, exist_ok=True)
                
                # Guardar archivo temporalmente
                filename = secure_filename(file.filename)
                filepath = os.path.join(UPLOAD_FOLDER, filename)
                file.save(filepath)
                
                # Procesar según extensión
                extension = filename.rsplit('.', 1)[1].lower()
                
                estudiantes_creados = 0
                calificaciones_importadas = 0
                errores = []
                
                if extension == 'csv':
                    # Procesar CSV
                    with open(filepath, 'r', encoding='utf-8') as csvfile:
                        reader = csv.DictReader(csvfile)
                        
                        for row in reader:
                            try:
                                nombre = row.get('Nombre', '').strip()
                                prueba = row.get('Prueba', '').strip()
                                nota = row.get('Nota', '').strip()
                                
                                if not nombre or not prueba or not nota:
                                    continue
                                
                                # Buscar o crear estudiante
                                estudiante = Estudiante.query.filter_by(
                                    id_asignacion=id_asignacion,
                                    nombre_alumno=nombre
                                ).first()
                                
                                if not estudiante:
                                    estudiante = Estudiante(
                                        id_asignacion=id_asignacion,
                                        nombre_alumno=nombre,
                                        estado_alumno='activo'
                                    )
                                    db.session.add(estudiante)
                                    db.session.flush()  # Para obtener el ID
                                    estudiantes_creados += 1
                                
                                # Crear calificación
                                calificacion = Calificacion(
                                    id_estudiante=estudiante.id_estudiante,
                                    nombre_prueba=prueba,
                                    nota_final=float(nota),
                                    fecha_importacion=datetime.now()
                                )
                                db.session.add(calificacion)
                                calificaciones_importadas += 1
                                
                            except Exception as e:
                                errores.append(f"Error en fila con {nombre}: {str(e)}")
                
                elif extension in ['xlsx', 'xls']:
                    # Procesar Excel
                    wb = openpyxl.load_workbook(filepath)
                    ws = wb.active
                    
                    # Asumir que la primera fila son encabezados
                    headers = [cell.value for cell in ws[1]]
                    
                    # Encontrar índices de columnas
                    nombre_idx = headers.index('Nombre') if 'Nombre' in headers else 0
                    prueba_idx = headers.index('Prueba') if 'Prueba' in headers else 1
                    nota_idx = headers.index('Nota') if 'Nota' in headers else 2
                    
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        try:
                            nombre = str(row[nombre_idx]).strip() if row[nombre_idx] else ''
                            prueba = str(row[prueba_idx]).strip() if row[prueba_idx] else ''
                            nota = row[nota_idx]
                            
                            if not nombre or not prueba or nota is None:
                                continue
                            
                            # Buscar o crear estudiante
                            estudiante = Estudiante.query.filter_by(
                                id_asignacion=id_asignacion,
                                nombre_alumno=nombre
                            ).first()
                            
                            if not estudiante:
                                estudiante = Estudiante(
                                    id_asignacion=id_asignacion,
                                    nombre_alumno=nombre,
                                    estado_alumno='activo'
                                )
                                db.session.add(estudiante)
                                db.session.flush()
                                estudiantes_creados += 1
                            
                            # Crear calificación
                            calificacion = Calificacion(
                                id_estudiante=estudiante.id_estudiante,
                                nombre_prueba=prueba,
                                nota_final=float(nota),
                                fecha_importacion=datetime.now()
                            )
                            db.session.add(calificacion)
                            calificaciones_importadas += 1
                            
                        except Exception as e:
                            errores.append(f"Error en fila: {str(e)}")
                
                # Confirmar cambios
                db.session.commit()
                
                # Eliminar archivo temporal
                os.remove(filepath)
                
                # Mostrar resultado
                mensaje = f'Importación completada: {calificaciones_importadas} calificaciones importadas'
                if estudiantes_creados > 0:
                    mensaje += f', {estudiantes_creados} estudiantes nuevos creados'
                
                flash(mensaje, 'success')
                
                if errores:
                    for error in errores[:5]:  # Mostrar máximo 5 errores
                        flash(error, 'warning')
                
                return redirect(url_for('calificaciones.listar', id_asignacion=id_asignacion))
                
            except Exception as e:
                db.session.rollback()
                if os.path.exists(filepath):
                    os.remove(filepath)
                flash(f'Error al procesar el archivo: {str(e)}', 'error')
                return redirect(request.url)
        
        else:
            flash('Tipo de archivo no permitido. Usa CSV o Excel (.xlsx, .xls)', 'error')
            return redirect(request.url)
    
    return render_template('calificaciones/importar.html', asignacion=asignacion)
